"""
Module d'export Excel avec gestion d'erreurs robuste
labCE v5.0
"""
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment
import os
import logging
import re
from datetime import datetime


def safe_sheet_title(raw_title):
    """Retourne un titre de feuille Excel valide."""
    title = re.sub(r'[:\\\/\?\*\[\]]', ' ', str(raw_title))
    title = re.sub(r'\s+', ' ', title).strip()
    title = title[:31]
    if not title:
        title = "Feuil1"
    return title


class ExcelExporter:
    """Gestionnaire d'export Excel complet."""

    def __init__(self, test_name, date_str, mode_name, graph_title=None):
        self.test_name = test_name
        self.date_str = date_str
        self.mode_name = mode_name
        self.graph_title = graph_title or test_name
        self.wb = openpyxl.Workbook()
        self.ws_main = self.wb.active
        self.ws_main.title = safe_sheet_title(self.graph_title)

        # Styles
        self.header_font = Font(name='Segoe UI', bold=True, size=10, color='FFFFFF')
        self.header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
        self.title_font = Font(name='Segoe UI', bold=True, size=13)

    def _write_header_row(self, row_data):
        """Écrit une ligne d'en-tête stylisée."""
        self.ws_main.append(row_data)
        for cell in self.ws_main[self.ws_main.max_row]:
            if cell.value:
                cell.font = self.header_font
                cell.fill = self.header_fill

    def add_header(self, curve_count=1):
        """Ajoute l'en-tête."""
        title = self.graph_title
        if curve_count > 1:
            title = f"{title} [{curve_count} courbes]"

        self.ws_main.append([title])
        self.ws_main['A1'].font = self.title_font
        self.ws_main.append(["Date", self.date_str])
        self.ws_main.append(["Mode", self.mode_name])
        self.ws_main.append([])

    def add_statistics(self, stats, durations=None):
        """Ajoute les statistiques."""
        if not stats:
            return
        self._write_header_row(["", "Min", "Max", "Moyenne", "Écart-type"])
        self.ws_main.append([
            "Effort (N)",
            stats['effort']['min'], stats['effort']['max'],
            stats['effort']['mean'], stats['effort']['std']
        ])
        self.ws_main.append([
            "Course (mm)",
            stats['course']['min'], stats['course']['max'],
            stats['course']['mean'], stats['course']['std']
        ])
        self.ws_main.append([])
        # Durées de chaque courbe
        if durations:
            self._write_header_row(["Courbe", "Durée (s)"])
            for name, dur in durations.items():
                self.ws_main.append([name, round(dur, 2)])
            self.ws_main.append([])

    def add_main_data(self, data):
        """Ajoute les données principales."""
        self._write_header_row([
            "Temps (s)", "Course (mm)", "Effort (N)",
            "Tension Effort (V)", "Tension Course (V)"
        ])
        for t, c, e, v_e, v_c in zip(
            data['time'], data['course'], data['effort'],
            data['raw_effort'], data['raw_course']
        ):
            self.ws_main.append([t, c, e, v_e, v_c])
        self.ws_main.append([])

    def add_points(self, points):
        """Ajoute les points marqués dans la feuille principale ET dans une feuille dédiée."""
        if not points:
            return

        # === Feuille principale (comme avant) ===
        self.ws_main.append(["POINTS DE MESURE"])
        self.ws_main[self.ws_main.max_row][0].font = Font(bold=True, size=11)
        self._write_header_row(["Nom", "X", "Y", "Type", "Courbe"])

        for point in points:
            if isinstance(point, dict):
                name = point.get('name', '')
                x = point.get('x', 0)
                y = point.get('y', 0)
                ptype = point.get('type', 'course')
                curve = point.get('curve', 'Principale')
                if curve is None:
                    curve = "Principale"
            elif isinstance(point, (list, tuple)) and len(point) >= 2:
                name = ""
                x, y = point[0], point[1]
                ptype = "course"
                curve = "Principale"
            else:
                continue
            self.ws_main.append([name, x, y, ptype, curve])

        self.ws_main.append([])

    def add_additional_curves(self, curves):
        """Ajoute les courbes importées."""
        for i, (name, crv) in enumerate(curves.items()):
            version_label = f"V{i+2}"  # V1 = principale, V2+ = importées
            self.ws_main.append([f"Courbe importée : {name} ({version_label})"])
            self.ws_main[self.ws_main.max_row][0].font = Font(bold=True, size=11)

            if "time" in crv and "effort" in crv:
                self._write_header_row(["Temps (s)", "Effort (N)"])
                for tt, ee in zip(crv["time"], crv["effort"]):
                    self.ws_main.append([tt, ee])
            elif "time" in crv and "course" in crv:
                self._write_header_row(["Temps (s)", "Course (mm)"])
                for tt, cc in zip(crv["time"], crv["course"]):
                    self.ws_main.append([tt, cc])
            elif "course" in crv and "effort" in crv:
                self._write_header_row(["Course (mm)", "Effort (N)"])
                for cc, ee in zip(crv["course"], crv["effort"]):
                    self.ws_main.append([cc, ee])

            self.ws_main.append([])

    def create_effort_course_sheet(self, data, curves, points=None):
        """Crée une feuille dédiée Course/Effort avec graphique.
        Col A = Course (mm), Col B = Effort (N) pour cohérence avec labCE.
        Les points de mesure sont dessinés comme des triangles sur le graphe."""
        try:
            ws_xy = self.wb.create_sheet("Effort_Course_XY")
            ws_xy.append(["Course (mm)", "Effort (N)"])

            pairs_main = sorted(
                zip(data['course'], data['effort']),
                key=lambda p: p[0]
            )
            for c_val, e_val in pairs_main:
                ws_xy.append([c_val, e_val])

            chart = ScatterChart()
            chart.title = self.graph_title
            chart.style = 2
            chart.x_axis.title = "Course (mm)"
            chart.y_axis.title = "Effort (N)"
            chart.width = 20
            chart.height = 14

            data_start = 2
            data_end = 1 + len(pairs_main)
            x_values = Reference(ws_xy, min_col=1, min_row=data_start, max_row=data_end)
            y_values = Reference(ws_xy, min_col=2, min_row=data_start, max_row=data_end)

            series_main = Series(y_values, x_values, title="V1 - Essai principal")
            series_main.marker.symbol = "circle"
            series_main.marker.size = 5
            chart.series.append(series_main)

            for i, (name, crv) in enumerate(curves.items()):
                if "effort" in crv and "course" in crv:
                    ws_xy.append([])
                    ws_xy.append([f"V{i+2} - {name}"])
                    row_start = ws_xy.max_row + 1
                    pairs = sorted(zip(crv["course"], crv["effort"]),
                                 key=lambda p: p[0])
                    for c_val, e_val in pairs:
                        ws_xy.append([c_val, e_val])
                    row_end = ws_xy.max_row
                    x_imp = Reference(ws_xy, min_col=1, min_row=row_start, max_row=row_end)
                    y_imp = Reference(ws_xy, min_col=2, min_row=row_start, max_row=row_end)
                    s = Series(y_imp, x_imp, title=f"V{i+2} - {name}")
                    s.marker.symbol = "diamond"
                    s.marker.size = 4
                    chart.series.append(s)

            # === Points de mesure sur le graphe (triangles) ===
            xy_points = []
            if points:
                for p in points:
                    if isinstance(p, dict):
                        px = p.get('x', 0)
                        py = p.get('y', 0)
                        pname = p.get('name', '')
                        xy_points.append((px, py, pname))
            if xy_points:
                ws_xy.append([])
                ws_xy.append(["Points - Course (mm)", "Points - Effort (N)"])
                pts_header_row = ws_xy.max_row
                pts_start = pts_header_row + 1
                for px, py, pn in xy_points:
                    ws_xy.append([px, py])
                pts_end = ws_xy.max_row
                x_pts = Reference(ws_xy, min_col=1, min_row=pts_start, max_row=pts_end)
                y_pts = Reference(ws_xy, min_col=2, min_row=pts_start, max_row=pts_end)
                s_pts = Series(y_pts, x_pts, title="Points de mesure")
                s_pts.marker.symbol = "triangle"
                s_pts.marker.size = 10
                s_pts.graphicalProperties.line.noFill = True  # pas de ligne, juste les marqueurs
                chart.series.append(s_pts)

            # === Tableau des points à côté du graphe (colonne H) ===
            if xy_points:
                pt_col = 8  # colonne H
                ws_xy.cell(row=1, column=pt_col, value="Nom").font = Font(bold=True)
                ws_xy.cell(row=1, column=pt_col+1, value="Course (mm)").font = Font(bold=True)
                ws_xy.cell(row=1, column=pt_col+2, value="Effort (N)").font = Font(bold=True)
                for ri, (px, py, pn) in enumerate(xy_points, start=2):
                    ws_xy.cell(row=ri, column=pt_col, value=pn)
                    ws_xy.cell(row=ri, column=pt_col+1, value=round(px, 3))
                    ws_xy.cell(row=ri, column=pt_col+2, value=round(py, 3))
                ws_xy.column_dimensions['H'].width = 12
                ws_xy.column_dimensions['I'].width = 14
                ws_xy.column_dimensions['J'].width = 14

            self.ws_main.add_chart(chart, "G5")

        except Exception as e:
            logging.error(f"Erreur création feuille Effort/Course : {e}")

    def create_time_graphs_sheet(self, data):
        """Crée une feuille avec graphiques en fonction du temps."""
        try:
            ws_extra = self.wb.create_sheet("Graph_Temps")
            ws_extra.append(["Temps (s)", "Effort (N)", "Course (mm)"])
            for t, e, c in zip(data['time'], data['effort'], data['course']):
                ws_extra.append([t, e, c])

            chart_et = ScatterChart()
            chart_et.title = "Effort = f(Temps)"
            chart_et.x_axis.title = "Temps (s)"
            chart_et.y_axis.title = "Effort (N)"
            t_ref = Reference(ws_extra, min_col=1, min_row=2, max_row=ws_extra.max_row)
            e_ref = Reference(ws_extra, min_col=2, min_row=2, max_row=ws_extra.max_row)
            s_et = Series(e_ref, t_ref, title="Effort")
            s_et.marker.symbol = "circle"
            chart_et.series.append(s_et)
            ws_extra.add_chart(chart_et, "E5")

            chart_ct = ScatterChart()
            chart_ct.title = "Course = f(Temps)"
            chart_ct.x_axis.title = "Temps (s)"
            chart_ct.y_axis.title = "Course (mm)"
            c_ref = Reference(ws_extra, min_col=3, min_row=2, max_row=ws_extra.max_row)
            s_ct = Series(c_ref, t_ref, title="Course")
            s_ct.marker.symbol = "triangle"
            chart_ct.series.append(s_ct)
            ws_extra.add_chart(chart_ct, "E25")

        except Exception as e:
            logging.error(f"Erreur création feuille temps : {e}")

    def save(self, filepath):
        """Sauvegarde le fichier Excel."""
        try:
            self.wb.save(filepath)
            logging.info(f"Fichier Excel sauvegardé : {filepath}")
            return True
        except PermissionError:
            logging.error(f"Permission refusée pour écrire : {filepath}")
            return False
        except Exception as e:
            logging.error(f"Erreur sauvegarde Excel : {e}")
            return False


def export_complete(data_manager, save_directory, test_name,
                   date_str, mode_name, graph_title=None):
    """
    Export complet avec toutes les données.

    Returns:
        tuple: (success, filepath)
    """
    try:
        if data_manager.is_empty and not data_manager.get_curve_names():
            logging.warning("Aucune donnée à exporter")
            return (False, None)

        if not os.path.isdir(save_directory):
            logging.error(f"Dossier invalide : {save_directory}")
            return (False, None)

        # Récupérer les données
        data = data_manager.get_data_copy()
        stats = data_manager.get_statistics()
        curve_count = 1 + len(data['additional_curves'])

        # Titre du fichier
        file_title = graph_title or test_name
        # Nom fichier avec nombre de courbes
        if curve_count > 1:
            filename = f"{test_name}_{curve_count}courbes.xlsx"
        else:
            filename = f"{test_name}.xlsx"

        filepath = os.path.join(save_directory, filename)

        # Calculer les durées de chaque courbe
        durations = {}
        if data['time'] and len(data['time']) > 1:
            durations["Principale"] = data['time'][-1] - data['time'][0]
        for name, crv in data['additional_curves'].items():
            if 'time' in crv and len(crv['time']) > 1:
                durations[name] = crv['time'][-1] - crv['time'][0]

        # Créer l'exporteur
        exporter = ExcelExporter(test_name, date_str, mode_name,
                                graph_title=graph_title)

        exporter.add_header(curve_count)
        exporter.add_statistics(stats, durations=durations)

        if not data_manager.is_empty:
            exporter.add_main_data(data)

        exporter.add_points(data['points'])
        exporter.add_additional_curves(data['additional_curves'])

        if not data_manager.is_empty:
            exporter.create_effort_course_sheet(data, data['additional_curves'], points=data['points'])
            exporter.create_time_graphs_sheet(data)

        success = exporter.save(filepath)
        return (success, filepath if success else None)

    except Exception as e:
        logging.exception("Erreur export complet")
        return (False, None)